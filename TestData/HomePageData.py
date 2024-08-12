import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstname":"Natan","email":"natan@paiva","password":"12345","gender":"Male"}, {"firstname":"Tania", "lastname":"tania@paiva", "password":"12345", "gender":"Female"}]

    @staticmethod
    def getTestData(test_case_name):
        data = []
        book = openpyxl.load_workbook("C:\\Users\\Natan\\Downloads\\homePageData.xlsx")
        sheet = book.active
        for i in range(2, sheet.max_row + 1):  # to get rows
            dict = {}
            #if sheet.cell(row=i, column=1).value == test_case_name:
            for j in range(2, sheet.max_column + 1):  # to get columns
                # Dict["email"]="natan@test.com
                dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            data.append(dict)
        return data

