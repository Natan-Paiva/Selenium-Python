import openpyxl

book = openpyxl.load_workbook('C:\\Users\\Natan\\Downloads\\homePageData.xlsx')
sheet = book.active
dict = {}
cell = sheet.cell(row=1,column=2)
print(cell.value)

#sheet.cell(row=2, column=2).value = 'Natan'

print(sheet.cell(row=2,column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['B2'].value)

for i in range(2,sheet.max_row+1): #rows
    if sheet.cell(row=i, column=1).value == "testCase2":
        for j in range(2, sheet.max_column+1): #columns
            dict[sheet.cell(row=1, column=j).value] = (sheet.cell(row=i, column= j).value)
            #print(sheet.cell(row=i, column= j).value)

print(dict)