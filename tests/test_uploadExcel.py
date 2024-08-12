import logging

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

from pageObjects.TablePage import TablePage
from utilities.BaseClass import BaseClass


class TestTablePage():
    def test_uploadoWithExel(self):

        def updateExcelData(searchTerm, colName, newValue):
            book = openpyxl.load_workbook('C:\\Users\\Natan\\Downloads\\download.xlsx')
            sheet = book.active
            dict = {}

            for i in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=i).value == colName:
                    dict["col"] = i

            for i in range(1, sheet.max_row + 1):
                for j in range(1, sheet.max_column + 1):
                    if sheet.cell(row=i, column=j).value == searchTerm:
                        dict["row"] = i

            sheet.cell(row=dict['row'], column=dict['col']).value = newValue
            book.save(file_path)

        file_path = 'C:/Users/Natan/Downloads/download.xlsx'
        fruitName = 'Apple'
        newValue = '999'
        driver = webdriver.Chrome()
        driver.implicitly_wait(5)
        driver.get('https://rahulshettyacademy.com/upload-download-test/')


        # download
        driver.find_element(By.CSS_SELECTOR, '#downloadButton').click()

        # edit excel
        updateExcelData('Apple', "price", newValue)

        # upload
        file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
        file_input.send_keys(file_path)

        wait = WebDriverWait(driver, 10)
        toast_locator = (By.XPATH, "//div[@class='Toastify__toast-body']/div[2]")
        wait.until(expected_conditions.visibility_of_element_located(toast_locator))
        msg = driver.find_element(*toast_locator).text

        assert 'Successfully' in msg

        priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
        value = driver.find_element(By.XPATH,
                                    "//div[text()='" + fruitName + "']/parent::div/parent::div/div[@id='cell-" + priceColumn + "-undefined']").text

        assert value == newValue